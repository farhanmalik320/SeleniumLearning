from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.keys import Keys
import time
import openpyxl as o
from selenium.webdriver.chrome.options import Options as ChromeOptions
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

excel_file = "C:\\Users\\cva\\PycharmProjects\\SeleniumLearning\\Excel\\Testfile.xlsx"
excel_worksheet= "Sheet1"
wb= o.load_workbook(excel_file)
ws= wb[excel_worksheet]
row_number= ws.max_row
col_number= ws.max_column

driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()))


def setURL():
    driver.get("https://gamesfi.live/")
    driver.maximize_window()

def login():

    login_btn = driver.find_element(By.XPATH, "//button[contains(text(),'LOGIN')]").click()

    # Skip the first row (headers)
    next(ws.rows)
    for row in ws.iter_rows():
        email = row[0].value
        password = row[1].value
        #print(email)
        #print(password)

        #Enter on email field
        email_field= driver.find_element(By.NAME,'email')
        email_field.send_keys(email)

        #Enter password field
        password_field= driver.find_element(By.NAME,'password')
        password_field.send_keys(password)

        #click on login button
        login_btn = driver.find_element(By.XPATH, "//button[contains(text(),'Login')]").click()

        time.sleep(2)

        email_field.clear()
        password_field.clear()

        actualerrormessage= driver.find_elements(By.CSS_SELECTOR, ".buxGtE")
        for error in actualerrormessage:
            print(error.text)

        time.sleep(2)

setURL()
login()