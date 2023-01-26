import openpyxl as o
excel_file = "C:\\Users\\cva\\PycharmProjects\\SeleniumLearning\\Excel\\Testfile.xlsx"
excel_worksheet= "Sheet1"
wb= o.load_workbook(excel_file)
ws= wb[excel_worksheet]
row_number= ws.max_row
col_number= ws.max_column
print("No of rows" , row_number)
print("No of col",  col_number)

row=2
print(ws.cell(row,1).value)
print(ws.cell(row,2).value)